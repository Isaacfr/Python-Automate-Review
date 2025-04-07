import openpyxl
from openpyxl import Workbook, load_workbook
import filenames

# Function to evaluate conditional formatting based on the specified condition
def evaluate_conditional_formatting(row):
    # Get the values from column B and D (assuming 0-based index)
    value_b = row[1].value  # Column B (index 1)
    value_d = row[3].value  # Column D (index 3)
    
    # Check if both values are not None and if they are trimmed and not equal
    return value_b is not None and value_d is not None and value_b.replace(" ","").lower() != value_d.replace(" ","").lower()

def filter_blanks_and_wrong_ids_function(src_file, src_sheet, lookup_file_path, lookup_table_file, lookup_full_path, lookup_table_sheet_name, lookup_value_col):
   workbook = openpyxl.load_workbook(src_file, data_only=True)
   source_sheet = workbook[src_sheet]  # Replace with your actual sheet name

   # Create a new worksheet for the results
   new_sheet = workbook.create_sheet(title='Wrong IDs')

   # Copy the header row to the new sheet
   for col_index, cell in enumerate(source_sheet[1], start=1):  # Assuming the first row is the header
       new_sheet.cell(row=1, column=col_index, value=cell.value)

   # Initialize a row index for the new sheet
   new_row_index = 2

   # Iterate through the rows in the source sheet
   for row in source_sheet.iter_rows(min_row=2):  # Start from row 2 to skip headers
       # Check if any cell in the row has a value of None, 0, or meets the conditional formatting condition
       if any(cell.value is None or cell.value == 0 for cell in row) or evaluate_conditional_formatting(row):
          # Copy the entire row to the new sheet
          for col_index, cell in enumerate(row, start=1):
             new_sheet.cell(row=new_row_index, column=col_index, value=cell.value)
          new_row_index += 1  # Move to the next row in the new sheet  

   # Determine the column where you want to add the new formula
   # For example, if you want to add a new column after the last column with data
   last_column = source_sheet.max_column
   new_column_index = last_column + 1  # New column index

   # Insert the header for the new column
   new_sheet.cell(row=1, column=new_column_index, value='Closest ID')

   #print(f'{lookup_full_path}')

   # Attach a new formula to the cells in the new column
   for row in range(2, new_sheet.max_row + 1):  # Start from row 2 to skip the header
      new_sheet.cell(row=row, column=new_column_index, value=f'=VLOOKUP(TRIM($B{row}),TRIM(\'{lookup_file_path}\\[{lookup_table_file}]{lookup_table_sheet_name}\'!$E:$G),lookup_value_col, FALSE)')

   # Save the workbook with the new sheet
   workbook.save(src_file)
   print(f"Rows with None, 0, or conditional formatting true have been written to the new sheet '{new_sheet.title}'.")