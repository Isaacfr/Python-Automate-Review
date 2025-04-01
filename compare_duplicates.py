import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import timedelta, date

date_dict = {}

def find_duplicates_and_export(input_file):
    # Load the Excel file
    df = pd.read_excel(input_file, sheet_name='Concise Statement')

    # Check for duplicates in the specified column (assuming 'ID' is the column to check)
    duplicates = df[df.duplicated(subset='BH Candidate ID', keep=False)]

    # Load the workbook and create a new sheet for duplicates
    with pd.ExcelWriter(input_file, engine='openpyxl', mode='a') as writer:
        duplicates.to_excel(writer, sheet_name='Duplicates', index=False)

def generate_date_range(start_date, end_date):
    """Generate a list of dates from start_date to end_date inclusive."""
    delta = end_date - start_date
    return [start_date + timedelta(days=i) for i in range(delta.days + 1)]

def create_date_ranges(input_file):
    # Load the workbook and select the active sheet
    workbook = load_workbook(input_file)
    sheet = workbook['Duplicates']

    """Load IDs and date ranges from an Excel file into the public date_dict variable."""
    global date_dict  # Declare that we are using the global variable

    # Iterate through the rows in the sheet (assuming data starts from row 2)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        id_value = row[0]  # Assuming ID is in the first column (A)
        start_date = row[1]  # Assuming Start Date is in the second column (B)
        end_date = row[2]  # Assuming End Date is in the third column (C)

        # Ensure start_date and end_date are of type date
        if isinstance(start_date, date) and isinstance(end_date, date):
            # Generate the date range
            date_range = generate_date_range(start_date, end_date)
        
            # Add to dictionary
            date_dict[id_value] = date_range

    # Print the resulting dictionary
    #for id_key, dates in date_dict.items():
        #print(f"ID: {id_key}, Dates: {dates}")

def create_id_mapping(file_path):
    """Create a new dictionary mapping one ID to another from an Excel file."""
    id_mapping = {}

    # Load the workbook and select the active sheet
    workbook = load_workbook(file_path)
    sheet = workbook['Duplicates']

    # Iterate through the rows in the sheet (assuming data starts from row 2)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key_id = row[7]  # Assuming the key ID is in the first column (A)
        value_id = row[0]  # Assuming the value ID is in the second column (B)

        # Check if the key already exists in the dictionary
        if key_id in id_mapping:
            # Append the value_id to the existing list
            id_mapping[key_id].append(value_id)
        else:
            # Create a new list with the value_id
            id_mapping[key_id] = [value_id]

    #for key_id, bookings in id_mapping.items():
        #print(f"ID: {key_id}, Bookings: {bookings}")

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Iterate through the id_mapping
    for key_id, value_ids in id_mapping.items():
        # Create a set to track dates
        date_occurrences = {}

        # Retrieve dates for each value ID and count occurrences
        for value_id in value_ids:
            if value_id in date_dict:
                for date in date_dict[value_id]:
                    if date in date_occurrences:
                        date_occurrences[date].append(value_id)
                    else:
                        date_occurrences[date] = [value_id]
        
        for date, value in date_occurrences.items():
            print(f"Dates: {date}, ID: {value}")

        # Highlight rows if any date occurs multiple times
        for date, ids in date_occurrences.items():
            if len(ids) > 1:  # If the date occurs more than once
                for row in sheet.iter_rows(min_row=2):
                    if row[0].value in ids:  # Assuming ID is in the first column (A)
                        for cell in row:
                            cell.fill = yellow_fill  # Highlight the entire row

    # Save the workbook after highlighting
    workbook.save(file_path)    
        
# Example usage
#input_file = r""

highlight_duplicates(source_file):
    find_duplicates_and_export(input_file)
    create_date_ranges(input_file)
    create_id_mapping(input_file)