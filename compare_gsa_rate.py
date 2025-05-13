from bs4 import BeautifulSoup
import requests
import pandas as pd
import filenames
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def compare_gsa_rate_function(year = 2025, zip_code = 32065, check_month='3'):

   #Utilize URL for web-scrapping
   url = f'https://www.gsa.gov/travel/plan-book/per-diem-rates/per-diem-rates-results?action=perdiems_report&fiscal_year={year}&state=&city=&zip={zip_code}'
   page = requests.get(url)
   soup = BeautifulSoup(page.text, features='html')

   #Create a default conversion map that converts a 3-letter date into a digit in string form
   month_mapping = {
       'jan': '1',
       'feb': '2',
       'mar': '3',
       'apr': '4',
       'may': '5',
       'jun': '6',
       'jul': '7',
       'aug': '8',
       'sep': '9',
       'oct': '10',
       'nov': '11',
       'dec': '12'
   }

   #Testing
   #print(soup)
   #soup.find('table')
   #soup.find_all('table')[1]

   #For this website, find the specific table and class to extract information
   table = soup.find('table', class_ = 'table_perdiem')
   if table is not None:
   
       #Testing to make sure identity is correct
       #table = soup.find('table', class_ = 'table_perdiem stripedTable dataTable no-footer dtr-inline')
       #table = soup.find_all('table')[1]
       #print(table)

       #Extract Headers
       world_titles = table.find_all('th')
       world_table_titles = [title.text.strip().replace('\xa0', '') for title in world_titles]
      
       #Testing values and what kind of data is being extracted
       #print(world_table_titles)
       #months = dict.fromkeys(world_table_titles[2:], None)
       months = world_table_titles[2:]
       new_months = []

       #Some data includes a year as well, this is to extract the 3-letter date
       for month in months:
             month_str = month[-3:].lower()
             new_months.append(month_mapping[month_str])

       #Testing difference
       #print(months)
       #print(new_months)

       #Create a dataframe object to extract data
       df = pd.DataFrame(columns = new_months)

       #Find all the rows and insert them into the dataframe
       #Create a blank dictionary for month data
       column_data = table.find_all('tr')
       months_dict = {}

       #Reiterate to each data and extract data, besides headers. 
       for row in column_data[1:]:
            row_data = row.find_all('td')
            individual_row_data = [data.text.strip() for data in row_data[2:]]
    
            #print(individual_row_data)
            length = len(df)
    
            df.loc[length] = individual_row_data

       row_index = 0
       months_dict = df.iloc[row_index].to_dict()

       #print(df)
       #print(months_dict)
   
       #Extract rate based on month that was used as input
       lodging_rate = months_dict[check_month]
       #print(lodging_rate)
       lodging_rate_value = float(lodging_rate.strip('$'))       

       return lodging_rate_value

   else:
       return 0

def highlight_values(src_file):
    wb = load_workbook(src_file)
    ws = wb['GSA Rate']

    # Define fill color for highlighting
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2):
        cell_a, cell_b = row[2], row[6]
        # Make sure both cells have numeric values before comparing
        if (cell_a.value is not None and cell_b.value is not None 
            and isinstance(cell_a.value, (int, float)) and isinstance(cell_b.value, (int, float))):
            if cell_a.value > cell_b.value:
                cell_a.fill = yellow_fill
                cell_b.fill = yellow_fill

    wb.save(src_file)


#highlight_values(filenames.source_file, filenames.source_sheet)

#Test calling the function
#compare_gsa_rate_function()