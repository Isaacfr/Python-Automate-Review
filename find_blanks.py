import openpyxl
import filenames

def find_blanks_function(file_path, source_sheet):

   wb = openpyxl.load_workbook(file_path)
   src_sheet = wb[source_sheet]

   # Create a new sheet
   new_sheet = wb.create_sheet(title='Blank Projects')  # Name the new sheet

   # Write headers in the new sheet (optional)
   for col in range(1,  src_sheet.max_column + 1):
       new_sheet.cell(row=1, column=col, value= src_sheet.cell(row=1, column=col).value)

   # Initialize a variable to keep track of the new row index
   new_row_index = 2  # Start from row 2 to leave space for headers

   # Iterate through the rows in the source sheet
   for row in  src_sheet.iter_rows(min_row=2):  # Start from row 2 to skip headers
       # Check if any cell in the row is blank or has a value of 0
       if any(cell.value is None or cell.value == 0 for cell in row):
           # Copy the entire row to the new sheet
           for col_index, cell in enumerate(row, start=1):
               new_sheet.cell(row=new_row_index, column=col_index, value=cell.value)
           new_row_index += 1  # Move to the next row in the new sheet

   # Save the changes to the Excel file
   wb.save(file_path)
   print(f"Data copied to {new_sheet} with blank entries in specified columns.")
