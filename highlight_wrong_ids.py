from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_wrong_ids_function(src_file, src_sheet):
    #Load workbook and select the active sheet  
    workbook = load_workbook(src_file)
    ws = workbook['Wrong IDs']

    bookings = []

    for cell in ws['A'][1:]:
         bookings.append(cell.value)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in workbook[src_sheet].iter_rows(min_row=2, min_col=1, max_col=1):
        if row[0].value in bookings:
            for cell in row:
                cell.fill = yellow_fill
        
    workbook.save(src_file)