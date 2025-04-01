import pandas as pd
from openpyxl import load_workbook
import filenames

def find_closest_id(update_file, update_sheet, lookup_file_path, lookup_table_file, lookup_full_path, lookup_table_sheet_name):
    workbook = load_workbook(update_file)
    sheet = workbook[update_sheet]
    sheet['E1'] = 'Closest ID'
    new_column_index = sheet.max_column + 1

    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=new_column_index, value=f'=VLOOKUP(TRIM($B{row}),TRIM(\'{lookup_file_path}\\[{lookup_table_file}]{lookup_table_sheet_name}\'!$E:$G),3, FALSE)')

    print(f'Included the closest id in {update_sheet}')

    workbook.save(update_file)

find_closest_id(filenames.source_file, 'Filtered Incorrect IDs', filenames.lookup_file_path, filenames.lookup_table_file, filenames.lookup_full_path, filenames.lookup_table_sheet_name)