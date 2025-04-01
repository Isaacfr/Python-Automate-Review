import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

def filter_ids_and_set_formula_function(source_file, src_sheet, lookup_file_path, lookup_table_file, lookup_full_path, lookup_table_sheet_name, lookup_value_col):

   wb = openpyxl.load_workbook(source_file)

   #Select the Source Sheet
   #source_sheet = wb['Sheet1']
   #source_sheet = wb['Concise Statement']
   source_sheet = wb[src_sheet]

   #Create a new Sheet
   destination_sheet = wb.create_sheet(title='Verify IDs')

   #Define Columns To Import
   columns_to_transfer = [1, 7, 8]  # Specify the columns to import

   # Initialize a variable to keep track of the new column index in the destination sheet
   new_col_index = 1

   # Copy the specified columns from the source sheet to the destination sheet
   for col_index in columns_to_transfer:
      for row_index in range(1, source_sheet.max_row + 1):  # Iterate through all rows
      #for row_index in range(1, 3):
         # Get the value from the source sheet
         cell_value = source_sheet.cell(row=row_index, column=col_index).value
         # Set the value in the destination sheet
         destination_sheet.cell(row=row_index, column=new_col_index, value=cell_value)
      new_col_index += 1  # Move to the next column in the destination sheet

   # Add a new column for XLOOKUP results
   xlookup_col_index = new_col_index  # This will be the next column after the transferred columns
   destination_sheet.cell(row=1, column=xlookup_col_index, value='XLOOKUP Result')  # Header for XLOOKUP results

   #Test if file opens correctly
   #try:
      #with open(lookup_table_file, 'r') as file:
         #print("The file path is valid and the file can be opened.")
   #except FileNotFoundError:
      #print("The file path is not valid.")
   #except Exception as e:
      #print(f"An error occurred: {e}")

   # Load the lookup table
   lookup_wb = openpyxl.load_workbook(lookup_full_path)
   lookup_sheet = lookup_wb[lookup_table_sheet_name]

   # Create an XLOOKUP formula for each row in the destination sheet
   for row_index in range(2, destination_sheet.max_row + 1):  # Start from row 2 to skip header
   #for row_index in range(2, 4):
      lookup_value = destination_sheet.cell(row=row_index, column=lookup_value_col).value
      if lookup_value is not None:  # Only create XLOOKUP if there is a lookup value
         # =TEXT(XLOOKUP(C2,'Temp ID Lookup'!$G:$G,'Temp ID Lookup'!$E:$E), 0)
         # Construct the XLOOKUP formula
         xlookup_formula = (
            f'=VALUETOTEXT('
    	    f'XLOOKUP(C{row_index}, '
            f'\'{lookup_file_path}\\[{lookup_table_file}]{lookup_table_sheet_name}\'!$G:$G, '
            f'\'{lookup_file_path}\\[{lookup_table_file}]{lookup_table_sheet_name}\'!$E:$E, 0))'
)
         destination_sheet.cell(row=row_index, column=xlookup_col_index, value=xlookup_formula)

   # Define the fill color for conditional formatting
   fill = PatternFill(start_color='FF7C80', end_color='FF7C80', fill_type='solid')
	 
   # Set conditional formatting for 2 columns to find different names from ID
   destination_sheet.conditional_formatting.add('$C2:$D4000',
     FormulaRule(formula=['TRIM($B2)<>TRIM($D2)'], fill=fill)
)

   # Save the changes to the Excel file
   wb.save(source_file)
   print(f"Please correct formula in {lookup_table_sheet_name}to proceed with next script")
